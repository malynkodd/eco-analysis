import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
from schemas import ReportInput

PRIMARY_HEX = "0F4C81"
GREEN_HEX   = "00B894"
ACCENT_HEX  = "6C5CE7"
LIGHT_BLUE  = "DBEAFE"
LIGHT_GREEN = "D1FAE5"
LIGHT_GRAY  = "F3F4F6"
WHITE       = "FFFFFF"


def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(cell, bg=PRIMARY_HEX):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_cell(cell, bold=False, center=False, bg=None, color=None):
    cell.font = Font(bold=bold, size=10, color=color or "1E293B")
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center"
    )
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def add_section_title(ws, row, title, color=PRIMARY_HEX, cols=7):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12, color=color)
    cell.fill = PatternFill("solid", fgColor="F8FAFC")
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=cols
    )
    return row + 1


def generate_excel(data: ReportInput) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ════════════════════════════════════════════════
    # ЛИСТ 1: Summary
    # ════════════════════════════════════════════════
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 40

    # Заголовок
    ws1["A1"] = "TECHNICO-ECONOMIC ANALYSIS REPORT"
    ws1["A1"].font = Font(bold=True, size=16, color=PRIMARY_HEX)
    ws1["A1"].fill = PatternFill("solid", fgColor="F0F4F8")
    ws1.merge_cells("A1:G1")
    ws1.row_dimensions[1].height = 30

    ws1["A2"] = "Environmental Measures Evaluation System"
    ws1["A2"].font = Font(size=11, color="64748B")
    ws1.merge_cells("A2:G2")
    ws1.row_dimensions[2].height = 20

    ws1.append([])

    # Інфо
    info = [
        ("Project:",           data.project_name),
        ("Description:",       data.project_description or "—"),
        ("Analyst:",           data.analyst_name),
        ("Generated:",         datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Measures analyzed:", str(len(data.financial_results))),
        ("Recommended:",       data.best_measure),
    ]
    for label, value in info:
        row = ws1.max_row + 1
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True, color=PRIMARY_HEX, size=10)
        cell = ws1.cell(row=row, column=2, value=value)
        cell.font = Font(size=10)
        if label == "Recommended:":
            cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
            cell.font = Font(bold=True, size=10, color="065F46")

    ws1.append([])
    ws1.append([])

    # Зведена таблиця
    add_section_title(ws1, ws1.max_row + 1, "📊 Consolidated Ranking", PRIMARY_HEX)
    headers = ["Measure", "NPV Rank", "CO2 Rank", "AHP Rank", "TOPSIS Rank", "Consensus"]
    ws1.append(headers)
    for cell in ws1[ws1.max_row]:
        style_header(cell, PRIMARY_HEX)

    for i, r in enumerate(sorted(data.ranking, key=lambda x: x.consensus_rank)):
        row_data = [
            r.name,
            f"#{r.rank_npv}",
            f"#{r.rank_co2}",
            f"#{r.rank_ahp}" if r.rank_ahp else "—",
            f"#{r.rank_topsis}" if r.rank_topsis else "—",
            f"#{r.consensus_rank}",
        ]
        ws1.append(row_data)
        bg = LIGHT_GREEN if r.consensus_rank == 1 else (LIGHT_BLUE if r.consensus_rank == 2 else None)
        for cell in ws1[ws1.max_row]:
            style_cell(cell, bold=(r.consensus_rank == 1), center=True, bg=bg)
        ws1[ws1.max_row][0].alignment = Alignment(horizontal="left", vertical="center")

    auto_width(ws1)

    # ════════════════════════════════════════════════
    # ЛИСТ 2: Financial Analysis
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("Financial Analysis")
    ws2.sheet_view.showGridLines = False

    add_section_title(ws2, 1, "💰 Financial Analysis — Key Indicators", PRIMARY_HEX)

    fin_headers = ["Measure", "NPV (UAH)", "IRR (%)", "BCR",
                   "Payback (yrs)", "Disc. Payback", "LCCA (UAH)"]
    ws2.append(fin_headers)
    for cell in ws2[ws2.max_row]:
        style_header(cell, PRIMARY_HEX)

    for f in data.financial_results:
        row_data = [
            f.name,
            f.npv,
            f"{f.irr:.1f}%",
            f"{f.bcr:.3f}",
            f"{f.simple_payback:.1f}" if f.simple_payback > 0 else "N/A",
            f"{f.discounted_payback:.0f}" if f.discounted_payback > 0 else "N/A",
            f.lcca,
        ]
        ws2.append(row_data)
        bg = LIGHT_GREEN if f.name == data.best_measure else None
        for cell in ws2[ws2.max_row]:
            style_cell(cell, bold=(f.name == data.best_measure), bg=bg)

        # NPV колір
        npv_cell = ws2.cell(row=ws2.max_row, column=2)
        npv_cell.number_format = '#,##0'
        if f.npv > 0:
            npv_cell.font = Font(bold=True, color="065F46", size=10)
        else:
            npv_cell.font = Font(bold=True, color="991B1B", size=10)

        # LCCA формат
        ws2.cell(row=ws2.max_row, column=7).number_format = '#,##0'

    ws2.append([])

    # Деталізація по роках
    if any(f.yearly_details for f in data.financial_results):
        add_section_title(ws2, ws2.max_row + 1,
                          "📈 Yearly Cash Flow Details", PRIMARY_HEX)

        yearly_headers = ["Measure", "Year", "Cash Flow (UAH)",
                          "Discounted CF", "Cumulative CF", "Cumulative Disc."]
        ws2.append(yearly_headers)
        for cell in ws2[ws2.max_row]:
            style_header(cell, ACCENT_HEX)

        for f in data.financial_results:
            if not f.yearly_details:
                continue
            for d in f.yearly_details:
                ws2.append([
                    f.name,
                    d.get("year"),
                    d.get("cash_flow"),
                    d.get("discounted_cash_flow"),
                    d.get("cumulative_cash_flow"),
                    d.get("cumulative_discounted"),
                ])
                for cell in ws2[ws2.max_row]:
                    style_cell(cell, center=(cell.column > 1))
                    if cell.column > 1:
                        cell.number_format = '#,##0'

    # Bar Chart NPV
    chart_start_row = 3
    chart = BarChart()
    chart.type = "col"
    chart.title = "NPV Comparison"
    chart.y_axis.title = "NPV (UAH)"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    names_col = 1
    values_col = 2
    data_rows = len(data.financial_results)

    data_ref = Reference(ws2, min_col=values_col, min_row=chart_start_row,
                         max_row=chart_start_row + data_rows)
    cats_ref = Reference(ws2, min_col=names_col, min_row=chart_start_row + 1,
                         max_row=chart_start_row + data_rows)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Розміщуємо chart справа від таблиці
    ws2.add_chart(chart, "I2")

    auto_width(ws2)

    # ════════════════════════════════════════════════
    # ЛИСТ 3: Environmental Impact
    # ════════════════════════════════════════════════
    ws3 = wb.create_sheet("Environmental Impact")
    ws3.sheet_view.showGridLines = False

    add_section_title(ws3, 1, "🌿 Environmental Impact Assessment", GREEN_HEX)

    eco_headers = ["Measure", "CO2 Reduction (t/yr)",
                   "Averted Damage (UAH)", "CO2 Value (USD)"]
    ws3.append(eco_headers)
    for cell in ws3[ws3.max_row]:
        style_header(cell, GREEN_HEX)

    for e in data.eco_results:
        ws3.append([
            e.name,
            e.co2_reduction_tons_per_year,
            e.averted_damage_uah,
            e.total_co2_value_usd,
        ])
        for cell in ws3[ws3.max_row]:
            style_cell(cell, center=(cell.column > 1))
        ws3[ws3.max_row][1].font = Font(bold=True, color="065F46", size=10)
        ws3[ws3.max_row][2].number_format = '#,##0'

    # Підсумок
    total_co2 = sum(e.co2_reduction_tons_per_year for e in data.eco_results)
    total_dmg = sum(e.averted_damage_uah for e in data.eco_results)
    ws3.append(["TOTAL", total_co2, total_dmg, "—"])
    for cell in ws3[ws3.max_row]:
        style_cell(cell, bold=True, center=(cell.column > 1), bg=LIGHT_GREEN)

    auto_width(ws3)

    # ════════════════════════════════════════════════
    # ЛИСТ 4: Sensitivity Analysis
    # ════════════════════════════════════════════════
    if data.sensitivity_data:
        ws4 = wb.create_sheet("Sensitivity Analysis")
        ws4.sheet_view.showGridLines = False

        add_section_title(ws4, 1, "🌪️ Sensitivity Analysis (Tornado)", PRIMARY_HEX)

        PARAM_LABELS = {
            "expected_savings":   "Expected Savings",
            "initial_investment": "Initial Investment",
            "discount_rate":      "Discount Rate",
            "operational_cost":   "Operational Cost",
            "lifetime_years":     "Lifetime (years)",
        }

        sens_headers = ["#", "Parameter", "Impact on NPV (UAH)", "Relative Impact (%)"]
        ws4.append(sens_headers)
        for cell in ws4[ws4.max_row]:
            style_header(cell, PRIMARY_HEX)

        max_impact = max(r.impact_percent for r in data.sensitivity_data) or 1
        for i, r in enumerate(data.sensitivity_data):
            rel = round(r.impact_percent / max_impact * 100, 1)
            ws4.append([
                i + 1,
                PARAM_LABELS.get(r.parameter, r.parameter),
                round(r.impact_percent),
                f"{rel}%",
            ])
            bg = LIGHT_BLUE if i == 0 else None
            for cell in ws4[ws4.max_row]:
                style_cell(cell, bold=(i == 0), center=(cell.column != 2), bg=bg)
            ws4[ws4.max_row][2].number_format = '#,##0'

        # Tornado bar chart
        tornado = BarChart()
        tornado.type = "bar"  # горизонтальний
        tornado.title = "Tornado Chart — Impact on NPV"
        tornado.x_axis.title = "Impact (UAH)"
        tornado.style = 10
        tornado.width = 20
        tornado.height = 14

        n = len(data.sensitivity_data)
        data_ref = Reference(ws4, min_col=3, min_row=2, max_row=2 + n)
        cats_ref = Reference(ws4, min_col=2, min_row=3, max_row=2 + n)
        tornado.add_data(data_ref, titles_from_data=True)
        tornado.set_categories(cats_ref)
        ws4.add_chart(tornado, "F2")

        auto_width(ws4)

    # ════════════════════════════════════════════════
    # ЛИСТ 5: AHP / TOPSIS
    # ════════════════════════════════════════════════
    if data.ahp_data or data.topsis_data:
        ws5 = wb.create_sheet("AHP & TOPSIS")
        ws5.sheet_view.showGridLines = False
        row = 1

        if data.ahp_data:
            row = add_section_title(ws5, row, "🎯 AHP — Criteria Weights", ACCENT_HEX)

            # CR статус
            cr = data.ahp_data.consistency_ratio
            cr_cell = ws5.cell(row=row, column=1,
                               value=f"Consistency Ratio (CR) = {cr:.4f} — "
                                     f"{'Consistent ✓' if cr < 0.1 else 'NOT Consistent ✗'}")
            cr_cell.font = Font(bold=True, size=10,
                                color="065F46" if cr < 0.1 else "991B1B")
            cr_cell.fill = PatternFill("solid",
                                       fgColor=LIGHT_GREEN if cr < 0.1 else "FEE2E2")
            ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

            ws5.append(["Criterion", "Weight", "Weight (%)"])
            for cell in ws5[ws5.max_row]:
                style_header(cell, ACCENT_HEX)
            row = ws5.max_row + 1

            max_w = max(data.ahp_data.weights)
            for c, w in zip(data.ahp_data.criteria, data.ahp_data.weights):
                ws5.append([c, round(w, 4), f"{w*100:.1f}%"])
                bg = LIGHT_BLUE if w == max_w else None
                for cell in ws5[ws5.max_row]:
                    style_cell(cell, bold=(w == max_w), bg=bg)

            ws5.append([])

            row = add_section_title(ws5, ws5.max_row + 1, "AHP Ranking", ACCENT_HEX)
            ws5.append(["Rank", "Alternative", "Weighted Score"])
            for cell in ws5[ws5.max_row]:
                style_header(cell, ACCENT_HEX)

            for r_item in data.ahp_data.ranking:
                ws5.append([f"#{r_item['rank']}", r_item["name"], round(r_item["score"], 4)])
                bg = LIGHT_GREEN if r_item["rank"] == 1 else None
                for cell in ws5[ws5.max_row]:
                    style_cell(cell, bold=(r_item["rank"] == 1), center=True, bg=bg)
                ws5[ws5.max_row][1].alignment = Alignment(horizontal="left")

            ws5.append([])

        if data.topsis_data:
            add_section_title(ws5, ws5.max_row + 1, "📐 TOPSIS Ranking", ACCENT_HEX)
            ws5.append(["Rank", "Alternative", "Closeness Coeff.",
                        "Distance to Ideal", "Distance to Anti-Ideal"])
            for cell in ws5[ws5.max_row]:
                style_header(cell, ACCENT_HEX)

            for r_item in data.topsis_data.ranking:
                ws5.append([
                    f"#{r_item['rank']}",
                    r_item["name"],
                    round(r_item["closeness_coefficient"], 4),
                    round(r_item["distance_to_ideal"], 4),
                    round(r_item["distance_to_anti_ideal"], 4),
                ])
                bg = LIGHT_GREEN if r_item["rank"] == 1 else None
                for cell in ws5[ws5.max_row]:
                    style_cell(cell, bold=(r_item["rank"] == 1), center=True, bg=bg)
                ws5[ws5.max_row][1].alignment = Alignment(horizontal="left")

        auto_width(ws5)

    # ════════════════════════════════════════════════
    # Зберігаємо
    # ════════════════════════════════════════════════
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()